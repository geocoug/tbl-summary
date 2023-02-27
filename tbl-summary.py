#!/usr/bin/env python

import argparse
import getpass
import logging
import os
import sys
from typing import Any

import openpyxl
import pandas as pd
import psycopg2
from openpyxl.styles import Border, Color, Font, PatternFill, Side
from openpyxl.utils.cell import get_column_letter

try:
    from _version import __vdate, __version__
except ImportError:
    __version__ = "unknown"
    __vdate = "unknown"


logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
formatter = logging.Formatter("%(message)s")
stream_handler = logging.StreamHandler()
stream_handler.setFormatter(formatter)
logger.addHandler(stream_handler)


def error_handler(error: Exception) -> None:
    """Simplify the output of all error messages."""
    err_type, err_obj, traceback = sys.exc_info()
    # line_num = traceback.tb_lineno
    logger.info(f"\n**** ERROR: {error}")
    sys.exit(1)


class Database:
    """Base database object."""

    def __init__(
        self: "Database",
        host: str,
        database: str,
        user: str,
    ) -> None:
        """Base database object.

        Args:
        ----
            host (str): Database host.
            database (str): Database name.
            user (str): Database user.
        """
        self.host = host
        self.database = database
        self.user = user
        self.port = 5432
        self.in_transaction = False
        self.encoding = "UTF8"
        self.conn = None
        self.get_password()

    def __repr__(self: "Database") -> str:
        """Pretty format the object."""
        return """{}( host={}, database={}, user={} )""".format(
            self.__class__.__name__,
            self.host,
            self.database,
            self.user,
        )

    def __del__(self: "Database") -> None:
        """Delete the instance."""
        self.close()

    def get_password(self: "Database") -> None:
        """Prompt the user to enter the database password."""
        self.passwd = getpass.getpass("Enter your password for %s: " % self.__repr__())
        try:
            self.open_db()
            self.close()
        except psycopg2.OperationalError as error:
            error_handler(error)
        finally:
            self.close()

    def open_db(self: "Database") -> None:
        """Open a database connection."""

        def db_conn(db: "Database"):  # noqa
            """Return a database connection object."""
            try:
                return psycopg2.connect(
                    host=str(db.host),
                    database=str(db.database),
                    port=db.port,
                    user=str(db.user),
                    password=str(db.passwd),
                )
            except psycopg2.OperationalError as err:
                raise err

        if self.conn is None:
            self.conn = db_conn(self)
        self.encoding = self.conn.encoding

    def cursor(self: "Database"):  # noqa
        """Return the connection cursor."""
        self.open_db()
        return self.conn.cursor()

    def close(self: "Database") -> None:
        """Close the database connection."""
        self.rollback()
        if self.conn is not None:
            self.conn.close()
            self.conn = None

    def commit(self: "Database") -> None:
        """Commit the current transaction."""
        if self.conn:
            self.conn.commit()
        self.in_transaction = False

    def rollback(self: "Database") -> None:
        """Roll back the current transaction."""
        if self.conn is not None:
            self.conn.rollback()
        self.in_transaction = False

    def execute(self: "Database", sql: str, params: dict = None):  # noqa
        """A shortcut to self.cursor().execute() that handles encoding.

        Handles insert, updates, deletes
        """
        self.in_transaction = True
        try:
            curs = self.cursor()
            if params is None:
                curs.execute(sql.encode(self.encoding))
            else:
                curs.execute(sql.encode(self.encoding), params)
        except Exception:
            self.rollback()
            raise
        return curs


class DataSummary(Database):
    """Methods used to create attributes which summarize a data table."""

    def __init__(
        self: "DataSummary",
        host: str,
        user: str,
        database: str,
        schema: str,
        table: str,
    ) -> None:
        """Methods used to create attributes which summarize a data table.

        Args:
        ----
            host (str): Database host.
            user (str): Database user.
            database (str): Database name.
            schema (str): Database schema.
            table (str): Database table.
        """
        super().__init__(host, database, user)
        self.schema = schema
        self.table = table
        self.verify_table_object()
        self.data = self.table_data()[0]
        self.total_rows = self.table_data()[1]
        self.df = pd.DataFrame(self.data, columns=self.columns())

    def verify_table_object(self: "DataSummary") -> None:
        """Check that the table object exists."""
        curs = self.execute(
            """
            SELECT table_schema, table_name
            FROM information_schema.tables
            WHERE table_schema = %(schema)s
            AND table_name = %(table)s;
            """,
            {"schema": self.schema, "table": self.table},
        )
        if curs.rowcount == 0:
            error_handler(
                Exception(f"Table object does not exist: {self.schema}.{self.table}"),
            )

    def columns(self: "DataSummary") -> list:
        """List all columns in the table."""
        cur = self.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = %(schema)s
            AND table_name = %(table)s;
            """,
            {"schema": self.schema, "table": self.table},
        )
        return self.column_rows(cur)

    def unique(self: "DataSummary", col: str) -> tuple[list, int]:
        """List all unique values in a column."""
        cur = self.execute(
            """
            SELECT distinct "%s"
            FROM %s.%s
            WHERE "%s" is not null
            ORDER BY "%s";
            """
            % (col, self.schema, self.table, col, col),
        )
        return self.column_rows(cur), cur.rowcount

    def column_rows(self: "DataSummary", cur: str) -> list:
        """Return a formatted list of records in the cursor.

        Args
        ----
            cur (str): Cursor object.

        Returns
        -------
            list: List of records.
        """
        return [v[0] for v in cur]

    def most_frequent_value(self: "DataSummary", col: str) -> tuple:
        """Select the most frequent values for a column in a table.

        Args
            col (str): Column to select most frequent values.

        Returns
        -------
            tuple: Tuple row object.
        """
        cur = self.execute(
            """
            SELECT "%s", COUNT(*)
            FROM %s.%s
            GROUP BY "%s"
            ORDER BY COUNT DESC;
            """
            % (col, self.schema, self.table, col),
        )
        return cur.fetchone()

    def column_dtype(self: "DataSummary", col: str) -> str:
        """Get the data type for a column in a table.

        Args
        ----
            col (str): Column name.

        Returns
        -------
            str: Data type for the column.
        """
        cur = self.execute(
            """SELECT
                    column_name,
                    case
                        when character_maximum_length is null
                        then data_type
                        else data_type || '(' || character_maximum_length || ')'
                        end as data_type
                FROM information_schema.columns
                WHERE table_schema = %(schema)s
                  AND table_name = %(table)s
                  AND column_name = %(col)s;
            """,
            {"col": col, "schema": self.schema, "table": self.table},
        )
        return cur.fetchone()[1]

    def table_data(self: "DataSummary") -> tuple[Any, int]:
        """Select all table data and number of rows in the table.

        Returns
        -------
            tuple: Tuple of data rows and the number of total rows.
        """
        try:
            cur = self.execute(
                """
                SELECT * FROM %s.%s;
                """
                % (self.schema, self.table),
            )
        except Exception as error:
            error_handler(error)
        return cur.fetchall(), cur.rowcount


def clparser() -> argparse.ArgumentParser:
    """Create a parser to handle input arguments."""
    desc_msg = f"""Create a summary of unique values for each column
        in a Postgres table. Summarize results in an Excel workbook.
        Version {__version__}, {__vdate}"""
    parser = argparse.ArgumentParser(description=desc_msg)
    parser.add_argument("output_file", help="Name of the XLSX output file")
    parser.add_argument(
        "-v",
        "--host",
        type=str,
        default="env3",
        dest="host",
        help="Server hostname.",
    )
    parser.add_argument(
        "-d",
        "--database",
        type=str,
        dest="database",
        help="Database name.",
    )
    parser.add_argument(
        "-s",
        "--schema",
        type=str,
        dest="schema",
        help="Database schema.",
    )
    parser.add_argument(
        "-u",
        "--username",
        type=str,
        dest="user",
        help="Database username.",
    )
    parser.add_argument(
        "-t",
        "--table",
        type=str,
        dest="table",
        help="Table name to summarize.",
    )
    return parser


def write_summary(data: "DataSummary", ofile: str) -> None:
    """Write results to an output file."""
    try:
        if os.path.exists(ofile):
            os.remove(ofile)
    except OSError as error:
        error_handler(error)

    wkb = openpyxl.Workbook()
    sheet = wkb.active
    sheet.title = "Data Summary"

    # Excel styling
    font = Font(bold=True, size=12)
    fill = PatternFill(patternType="solid", fgColor=Color("97b4c9"))
    border_all = Border(
        left=Side(border_style="thin", color=Color("000000")),
        right=Side(border_style="thin", color=Color("000000")),
        top=Side(border_style="thin", color=Color("000000")),
        bottom=Side(border_style="thin", color=Color("000000")),
    )

    # Summary header
    sheet.cell(row=1, column=2).value = "Host"
    sheet.cell(row=1, column=2).font = font
    sheet.cell(row=2, column=2).value = data.host

    sheet.cell(row=1, column=3).value = "Database"
    sheet.cell(row=1, column=3).font = font
    sheet.cell(row=2, column=3).value = data.database

    sheet.cell(row=1, column=4).value = "Schema"
    sheet.cell(row=1, column=4).font = font
    sheet.cell(row=2, column=4).value = data.schema

    sheet.cell(row=1, column=5).value = "Table"
    sheet.cell(row=1, column=5).font = font
    sheet.cell(row=2, column=5).value = data.table

    sheet.cell(row=1, column=6).value = "Total Rows"
    sheet.cell(row=1, column=6).font = font
    sheet.cell(row=2, column=6).value = data.total_rows

    sheet.cell(row=1, column=7).value = "Description"
    sheet.cell(row=1, column=7).font = font
    sheet.cell(row=2, column=7).value = "Distinct values for each column in a table."

    # Summary results
    for col in range(len(data.columns())):
        # column = data.columns()[col]
        if col == 0:
            sheet.cell(row=4, column=col + 1).value = "# of unique values"
            sheet.cell(row=4, column=col + 1).font = font
            sheet.cell(row=5, column=col + 1).value = "most frequent value"
            sheet.cell(row=5, column=col + 1).font = font
            sheet.cell(row=6, column=col + 1).value = "value frequency"
            sheet.cell(row=6, column=col + 1).font = font
            sheet.cell(row=7, column=col + 1).value = "data type"
            sheet.cell(row=7, column=col + 1).font = font
            sheet.cell(row=8, column=col + 1).value = "column name"
            sheet.cell(row=8, column=col + 1).font = font
        # Number of unique values
        sheet.cell(row=4, column=col + 2).value = data.unique(data.columns()[col])[1]
        # Most frequent value and frequency
        val, freq = data.most_frequent_value(data.columns()[col])
        sheet.cell(row=5, column=col + 2).value = val
        sheet.cell(row=6, column=col + 2).value = freq
        # Column data type
        sheet.cell(row=7, column=col + 2).value = data.column_dtype(data.columns()[col])
        # Column names
        sheet.cell(row=8, column=col + 2).value = data.columns()[col]
        sheet.cell(row=8, column=col + 2).font = font
        sheet.cell(row=8, column=col + 2).border = border_all
        sheet.cell(row=8, column=col + 2).fill = fill
        # Unique column values
        row = 9
        for value in data.unique(data.columns()[col])[0]:
            sheet.cell(row=row, column=col + 2).value = value
            row += 1
        # Turn on filter
        sheet.auto_filter.ref = f"B8:{get_column_letter(len(data.columns()) + 1)}8"
        # Column A cell width
        sheet.column_dimensions["A"].width = 25

    wkb.save(ofile)


def tbl_summary(
    host: str,
    user: str,
    database: str,
    schema: str,
    table: str,
    output_file: str,
) -> None:
    """Run the table summary on a PostgreSQL table.

    Args
    ----
        host (str): Database host.
        user (str): Database user.
        database (str): Database name.
        schema (str): Database schema.
        table (str): Database table.
        output_file (str): Excel file to save results.

    Raises
    ------
        Exception: Checks whether the provided output_file extension is xlxs
    """
    logger.info(f"{os.path.basename(__file__)}, {__version__}, {__vdate}")
    if os.path.splitext(output_file)[-1] not in [".xlsx"]:
        error_handler(Exception(f"File extension not valid: {output_file}"))

    data = DataSummary(host, user, database, schema, table)
    write_summary(data, output_file)
    logger.info(f"Results saved to {output_file}")


if __name__ == "__main__":
    parser = clparser()
    args = parser.parse_args()
    tbl_summary(**vars(args))
